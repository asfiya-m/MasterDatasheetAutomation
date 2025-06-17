from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from io import BytesIO

"""
populate_syscad_inputs_rev2.py

Accepts a **nested mapping**: {equipment_type: {master_param: stream_tag}}
so each equipment sheet can use its own tag names.
"""

def populate_syscad_inputs(master_file: BytesIO, streamtable_file: BytesIO, mapping_dict: dict):
    # Load workbooks
    master_wb = load_workbook(master_file)
    streamtable_wb = load_workbook(streamtable_file, data_only=True)

    # Styling
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    master_sheets = set(master_wb.sheetnames)
    stream_sheets = set(streamtable_wb.sheetnames)
    common_sheets = master_sheets & stream_sheets
    missing_sheets = list(master_sheets - stream_sheets)

    for eq_type in common_sheets:
        master_ws = master_wb[eq_type]
        stream_ws = streamtable_wb[eq_type]

        # -------- per‑equipment mapping ---------
        param_mapping = mapping_dict.get(eq_type, {})
        if not param_mapping:
            continue  # nothing mapped for this sheet

        # Unit tags (row 1, col D→)
        stream_unit_tags = [c.value for c in stream_ws[1][3:] if c.value]
        for i, tag in enumerate(stream_unit_tags):
            cell = master_ws.cell(row=3, column=4 + i, value=tag)
            cell.font = Font(bold=True)
            cell.border = thin

        # tag → row lookup (tag names assumed in column C / index 2)
        stream_tag_to_row = {
            row[2].value.strip(): r_idx
            for r_idx, row in enumerate(stream_ws.iter_rows(min_row=3, max_col=3), start=3)
            if row[2].value
        }

        # param → row lookup inside SysCAD Inputs block of master sheet
        param_rows, in_syscad = {}, False
        for r in range(1, master_ws.max_row + 1):
            label = master_ws.cell(row=r, column=1).value
            if label and "SysCAD Inputs" in str(label):
                in_syscad = True
            if in_syscad:
                if label and "Engineering Inputs" in str(label):
                    break
                pname = master_ws.cell(row=r, column=2).value
                if pname:
                    param_rows[pname.strip()] = r

        master_unit_tags = [c.value for c in master_ws[3][3:] if c.value]

        # -------- populate values --------
        for col_off, unit_tag in enumerate(master_unit_tags):
            if unit_tag not in stream_unit_tags:
                continue
            stream_col = stream_unit_tags.index(unit_tag) + 4  # Excel D=4
            for master_param, stream_tag in param_mapping.items():
                if master_param not in param_rows or stream_tag not in stream_tag_to_row:
                    continue
                m_row = param_rows[master_param]
                s_row = stream_tag_to_row[stream_tag]
                val = stream_ws.cell(row=s_row, column=stream_col).value
                if val is not None:
                    m_col = col_off + 4
                    cell = master_ws.cell(row=m_row, column=m_col, value=round(val, 2) if isinstance(val, float) else val)
                    cell.border = thin

                # update unit if differs
                stream_unit = stream_ws.cell(row=s_row, column=2).value  # column B
                master_unit = master_ws.cell(row=m_row, column=3).value
                if stream_unit and master_unit != stream_unit:
                    ucell = master_ws.cell(row=m_row, column=3, value=stream_unit)
                    ucell.border = thin

    buf = BytesIO()
    master_wb.save(buf)
    buf.seek(0)
    return buf, missing_sheets
