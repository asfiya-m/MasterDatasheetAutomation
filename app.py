"""

app.py

Streamlit frontend for automating the generation and population of a Master Equipment Datasheet.

Steps:
1. Upload a raw .xlsm file with multiple equipment sheets.
2. Generate a categorized master datasheet with grouped input sections.
3. Upload a SysCAD streamtable Excel file to populate SysCAD Inputs into the master datasheet.
4. Download final Excel file with all populated data.

Author: Asfiya Khanam
Created: June 2025

"""
import streamlit as st
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook

# Backend code
from automation_test1 import generate_master_datasheet  # Phase‑1 generator
from populate_syscad_inputs_rev2 import populate_syscad_inputs  # Phase‑2 backend (per‑equipment mapping)

# Page configuration
st.set_page_config(page_title="Master Datasheet Automation", page_icon="📄", layout="wide")

# --------------------------------------------------
# Shared Title and description
# --------------------------------------------------

st.title("📄 Master Equipment Datasheet Automation Tool")
st.markdown(
    """
This tool helps you:
1. Generate a clean, categorized master datasheet from your raw Excel input.
2. Match each SysCAD parameters in the Master Datasheet to the correct tag in your SysCAD streamtable.
3. Populate the Master Datasheet with SysCAD streamtable data.
"""
)

# --------------------------------------------------
# Sidebar navigation
# --------------------------------------------------

st.sidebar.title("📚 Menu")
page = st.sidebar.radio("Choose a page", ("Generate Master Datasheet", "SysCAD parameters- Map & Populate"))
# if "page" not in st.session_state:
#     st.session_state["page"] = "Generate Master Datasheet"

# page = st.sidebar.radio(
#     "Choose a page",
#     ("Generate Master Datasheet", "SysCAD parameters- Map & Populate"),
#     index=0 if st.session_state["page"] == "Generate Master Datasheet" else 1,
# )
# ==================================================
# PAGE 1 – GENERATE MASTER DATASHEET
# ==================================================

if page == "Generate Master Datasheet":
    st.header("Step 1 – Generating Master Datasheet")
    st.markdown(
        """
        **What happens in this step?**
        - The tool parses every equipment sheet and extracts equipment-wise parameters.
        - Groups the parameters under 5 categories:
            - SysCAD Inputs
            - Engineering Inputs
            - Lab/Pilot Inputs
            - Project Constants
            - Vendor Inputs
        - Creates one formatted sheet per equipment.

        **To do:**
        Upload the Datasheets workbook
        """
    )

    uploaded_raw = st.file_uploader("Upload original equipment datasheets workbook `.xlsm`", type=["xlsm"], key="raw_file")

    if uploaded_raw and st.button("Generate Master Sheet"):
        output_stream, outfile_name = generate_master_datasheet(BytesIO(uploaded_raw.read()))
        output_stream.seek(0)
        st.session_state["generated_master"] = output_stream
        st.success("Master datasheet generated ✔️")
        st.download_button(
            "📥 Download Master Sheet",
            data=output_stream,
            file_name=outfile_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    # st.markdown("---")
    # if st.button("➡️ Next: Go to Mapping and Population"):
    #     st.session_state["page"] = "SysCAD parameters- Map & Populate"
    #     st.rerun()  # Forces the app to update immediately
 

# ==================================================
# PAGE 2 – MAP PARAMETERS & POPULATE VALUES
# ==================================================

if page == "SysCAD parameters- Map & Populate":
    st.header("Step 2 – SysCAD Parameter Mapping & Data Population")

    # ---------- File selection ----------
    st.subheader("Select input files")
    col1, col2 = st.columns(2)
    with col1:
        master_option = st.radio(
            "Master sheet source",
            ("Use the generated Master Datasheet from Step 1", "Upload existing master `.xlsx``"),
            key="master_src_radio",
        )
    with col2:
        stream_upload = st.file_uploader("Upload SysCAD streamtable `.xlsx`", type=["xlsx"], key="stream_file")

    # Resolve master sheet bytes
    if master_option.startswith("Use"):
        master_bytes = st.session_state.get("generated_master")
    else:
        master_upload = st.file_uploader("Upload master sheet", type=["xlsx"], key="master_file")
        master_bytes = BytesIO(master_upload.read()) if master_upload else None

    if not master_bytes or not stream_upload:
        st.info("Please provide **both** master datasheet and streamtable to continue.")
        st.stop()

    stream_bytes = BytesIO(stream_upload.read())

    # ---------- Analyse workbooks ----------
    master_wb = load_workbook(master_bytes, read_only=True)
    stream_wb = load_workbook(stream_bytes, read_only=True, data_only=True)

    master_equipment = set(master_wb.sheetnames)
    stream_equipment = set(stream_wb.sheetnames)
    common_equipment = sorted(master_equipment & stream_equipment)
    missing_equipment = sorted(master_equipment - stream_equipment)

    if missing_equipment:
        st.warning("❗ Equipment sheets missing in streamtable that are available in the Master Datasheet: " + ", ".join(missing_equipment))

    # ---------- Helper to extract SysCAD parameters ----------
    def extract_syscad_params(ws):
        params, in_syscad = [], False
        for r in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=1).value
            if cell_val and "SysCAD Inputs" in str(cell_val):
                in_syscad = True
            if in_syscad:
                if cell_val and "Engineering Inputs" in str(cell_val):
                    break
                p = ws.cell(row=r, column=2).value
                if p:
                    params.append(p.strip())
        return params

    # ---------- Build mapping UI ----------
    st.subheader("Map parameters for each equipment type")
    st.markdown("Select a streamtable **tag** for every SysCAD Input parameter. Leave blank to skip.")

    if "tmp_mapping" not in st.session_state:
        st.session_state["tmp_mapping"] = {}

    for equip in common_equipment:
        master_ws = master_wb[equip]
        stream_ws = stream_wb[equip]

        param_list = extract_syscad_params(master_ws)
        tag_list = sorted({row[2].value.strip() for row in stream_ws.iter_rows(min_row=3, max_col=3) if row[2].value})

        if not param_list:
            st.info(f"⚠️  SysCAD parameters not found in **{equip}** Master Datasheet. Skipping.")
            continue
        if not tag_list:
            st.info(f"⚠️  Tags not available for **{equip}** in the SysCAD streamtable. Skipping.")
            continue

        st.subheader(equip)
        eq_map = st.session_state["tmp_mapping"].setdefault(equip, {})
        for p in param_list:
            default = eq_map.get(p, "— skip —")
            choice = st.selectbox(
                p,
                ["— skip —"] + tag_list,
                key=f"{equip}_{p}",
                index=(["— skip —"] + tag_list).index(default) if default in ["— skip —"] + tag_list else 0,
            )
            if choice == "— skip —":
                eq_map.pop(p, None)
            else:
                eq_map[p] = choice

    # ---------- Save mapping ----------
    if st.button("💾 Save Mapping"):
        if any(st.session_state["tmp_mapping"].values()):
            st.session_state["param_mapping"] = st.session_state["tmp_mapping"]
            st.success("Mapping saved! You can now populate the master sheet.")
        else:
            st.error("No mappings selected. Nothing saved.")

    st.markdown("---")
    st.subheader("Populate Master Datasheet with SysCAD values")

    if st.button("🚀 Populate & Download"):
        mapping_final = st.session_state.get("param_mapping")
        if not mapping_final:
            st.error("Please save a parameter mapping first.")
            st.stop()

        master_bytes.seek(0)
        stream_bytes.seek(0)

        # Pass nested mapping directly
        populated_stream, missing = populate_syscad_inputs(master_bytes, stream_bytes, mapping_final)

        if missing:
            st.warning("Streamtable missing for: " + ", ".join(missing))

        populated_stream.seek(0)
        fname = f"Master_DataSheet_SysCADPopulated_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        st.download_button(
            "📥 Download Populated Sheet",
            data=populated_stream,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
